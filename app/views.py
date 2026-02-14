from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .models.garantia import Garantia
from datetime import datetime


# Lista de piezas booleanas del modelo Garantia
PIEZAS_FIELDS = [
    ('tubo_escape', 'Tubo de escape'),
    ('tarjeta', 'Tarjeta'),
    ('motor_1000w', 'Motor de 1000w'),
    ('motor_1200w', 'Motor 1200w'),
    ('motor_1500w', 'Motor 1500w'),
    ('cargador_bateria', 'Cargador de batería'),
    ('baterias', 'Baterías'),
    ('caja_reguladora', 'Caja reguladora'),
    ('diferencial', 'Diferencial'),
    ('extensor_rango', 'Extensor de rango'),
    ('problema_electrico', 'Problema eléctrico'),
    ('caja_luces', 'Caja luces'),
    ('farol_delantero', 'Farol delantero'),
    ('farol_trasero', 'Farol trasero'),
    ('rodamientos_direccion', 'Rodamientos dirección'),
    ('rodamientos_delanteros', 'Rodamientos delanteros'),
    ('rodamientos_traseros', 'Rodamientos traseros'),
    ('bandas_freno', 'Bandas de freno'),
    ('claxon', 'Claxon'),
    ('pantalla', 'Pantalla'),
]


@staff_member_required
def reporte_garantias_view(request):
    """Vista para seleccionar el rango de fechas del reporte"""
    return render(request, 'app/reporte_garantias.html')


@staff_member_required
def generar_reporte_garantias_excel(request):
    """Genera el reporte Excel de garantías por piezas"""
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if not fecha_inicio or not fecha_fin:
        return HttpResponse("Debe proporcionar fecha_inicio y fecha_fin", status=400)
    
    try:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Formato de fecha inválido. Use YYYY-MM-DD", status=400)
    
    # Filtrar garantías por rango de fechas que tengan triciclo asociado
    garantias = Garantia.objects.filter(
        fecha_creacion__gte=fecha_inicio,
        fecha_creacion__lte=fecha_fin,
        triciclo__isnull=False
    ).select_related('triciclo')
    
    # Crear diccionario para almacenar VINs por cada pieza
    piezas_vins = {field_name: [] for field_name, _ in PIEZAS_FIELDS}
    otros_vins = []  # Para el campo "otros"
    
    for garantia in garantias:
        vin = garantia.triciclo.vin if garantia.triciclo else None
        if not vin:
            continue
            
        for field_name, _ in PIEZAS_FIELDS:
            if getattr(garantia, field_name, False):
                piezas_vins[field_name].append(vin)
        
        # Campo "otros" si tiene contenido
        if garantia.otros and garantia.otros.strip():
            otros_vins.append((vin, garantia.otros))
    
    # Crear el archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Garantías"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título del reporte
    ws.merge_cells('A1:' + get_column_letter(len(PIEZAS_FIELDS) + 1) + '1')
    ws['A1'] = f"Reporte de Garantías - VINs por Pieza ({fecha_inicio} al {fecha_fin})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Escribir encabezados (nombres de piezas)
    for col, (field_name, display_name) in enumerate(PIEZAS_FIELDS, start=1):
        cell = ws.cell(row=3, column=col, value=display_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = 22
    
    # Agregar columna "Otros"
    otros_col = len(PIEZAS_FIELDS) + 1
    cell = ws.cell(row=3, column=otros_col, value="Otros")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(otros_col)].width = 30
    
    # Encontrar el máximo número de filas necesarias
    max_rows = max(
        max((len(vins) for vins in piezas_vins.values()), default=0),
        len(otros_vins)
    )
    
    # Escribir los VINs
    for col, (field_name, _) in enumerate(PIEZAS_FIELDS, start=1):
        vins = piezas_vins[field_name]
        for row, vin in enumerate(vins, start=4):
            cell = ws.cell(row=row, column=col, value=vin)
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # Escribir "Otros" (VIN + descripción)
    for row, (vin, descripcion) in enumerate(otros_vins, start=4):
        cell = ws.cell(row=row, column=otros_col, value=f"{vin}: {descripcion}")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # Rellenar celdas vacías con bordes hasta max_rows
    if max_rows > 0:
        for col in range(1, len(PIEZAS_FIELDS) + 2):
            for row in range(4, max_rows + 4):
                cell = ws.cell(row=row, column=col)
                if cell.value is None:
                    cell.value = ""
                cell.border = thin_border
    
    # Agregar resumen al final
    summary_row = max(max_rows + 6, 6)
    ws.cell(row=summary_row, column=1, value="RESUMEN - Total de VINs por pieza:").font = Font(bold=True, size=12)
    
    summary_row += 1
    for col, (field_name, display_name) in enumerate(PIEZAS_FIELDS, start=1):
        count = len(piezas_vins[field_name])
        if count > 0:
            ws.cell(row=summary_row, column=1, value=display_name)
            ws.cell(row=summary_row, column=2, value=count)
            summary_row += 1
    
    if otros_vins:
        ws.cell(row=summary_row, column=1, value="Otros")
        ws.cell(row=summary_row, column=2, value=len(otros_vins))
    
    # Preparar la respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_garantias_{fecha_inicio}_{fecha_fin}.xlsx"'
    
    wb.save(response)
    return response
